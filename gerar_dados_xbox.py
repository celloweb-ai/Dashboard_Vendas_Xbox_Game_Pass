#!/usr/bin/env python3
"""
Script para gerar arquivos Excel do Dashboard de Vendas Xbox Game Pass
Autor: Marcus Vasconcellos
Data: Janeiro 2026
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configurar seed para reprodutibilidade
np.random.seed(42)
random.seed(42)

print("🎮 Gerando dados do Xbox Game Pass...")

# ===== GERAR BASE DE DADOS =====
subscriber_ids = range(3231, 3526)  # 295 registros
names_primeiro = ['Marcus', 'Ana', 'Carlos', 'Julia', 'Pedro', 'Maria', 'João', 'Fernanda', 'Lucas', 'Beatriz',
                  'Rafael', 'Camila', 'Felipe', 'Gabriela', 'Ricardo', 'Patricia', 'Gustavo', 'Mariana', 'Thiago', 'Amanda']
names_ultimo = ['Silva', 'Santos', 'Oliveira', 'Sousa', 'Costa', 'Martins', 'Ferreira', 'Gomes', 'Pereira', 'Dias',
                'Alves', 'Ribeiro', 'Monteiro', 'Barbosa', 'Cardoso', 'Rocha', 'Lopes', 'Nunes', 'Pinto', 'Correia']

# Gerar dados
data = {
    'Subscriber ID': list(subscriber_ids),
    'Name': [f"{random.choice(names_primeiro)} {random.choice(names_ultimo)}" for _ in range(295)],
    'Plan': np.random.choice(['Ultimate', 'Standard', 'Core'], 295, p=[0.35, 0.40, 0.25]),
    'Start Date': [datetime(2024, 1, 1) + timedelta(days=random.randint(0, 365)) for _ in range(295)],
    'Auto Renewal': np.random.choice(['Yes', 'No'], 295, p=[0.60, 0.40]),
    'Subscription Type': np.random.choice(['Monthly', 'Quarterly', 'Annual'], 295, p=[0.50, 0.25, 0.25]),
}

df = pd.DataFrame(data)

# Definir preços
plan_prices = {'Ultimate': 15, 'Standard': 10, 'Core': 5}
ea_play_price = 30
minecraft_price = 20

df['Subscription Price'] = df['Plan'].map(plan_prices)
df['EA Play Season Pass'] = df['Plan'].apply(lambda x: ea_play_price if x == 'Ultimate' and random.random() > 0.6 else 0)
df['Minecraft Season Pass'] = df['Plan'].apply(lambda x: minecraft_price if x in ['Ultimate', 'Standard'] and random.random() > 0.5 else 0)
df['Coupon Value'] = np.random.choice([0, 0, 0, 5, 10, 15, 20, 25], 295)
df['Total Value'] = (df['Subscription Price'] + df['EA Play Season Pass'] + df['Minecraft Season Pass'] - df['Coupon Value']).round(2)

# ===== CRIAR ARQUIVO base.xlsx =====
print("📊 Criando base.xlsx...")

wb_base = Workbook()
ws_base = wb_base.active
ws_base.title = "Bases"

# Cores Xbox
header_fill = PatternFill(start_color="9BC848", end_color="9BC848", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Escrever cabeçalhos
headers = ['Subscriber ID', 'Name', 'Plan', 'Start Date', 'Auto Renewal', 'Subscription Price',
           'Subscription Type', 'EA Play Season Pass', 'Minecraft Season Pass', 'Coupon Value', 'Total Value']

for col_num, header in enumerate(headers, 1):
    cell = ws_base.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Escrever dados
for row_num, (idx, row) in enumerate(df.iterrows(), 2):
    ws_base.cell(row=row_num, column=1, value=int(row['Subscriber ID']))
    ws_base.cell(row=row_num, column=2, value=row['Name'])
    ws_base.cell(row=row_num, column=3, value=row['Plan'])
    ws_base.cell(row=row_num, column=4, value=row['Start Date'].strftime('%d/%m/%Y'))
    ws_base.cell(row=row_num, column=5, value=row['Auto Renewal'])
    ws_base.cell(row=row_num, column=6, value=row['Subscription Price'])
    ws_base.cell(row=row_num, column=7, value=row['Subscription Type'])
    ws_base.cell(row=row_num, column=8, value=row['EA Play Season Pass'])
    ws_base.cell(row=row_num, column=9, value=row['Minecraft Season Pass'])
    ws_base.cell(row=row_num, column=10, value=row['Coupon Value'])
    ws_base.cell(row=row_num, column=11, value=row['Total Value'])
    
    # Aplicar bordas
    for col_num in range(1, 12):
        ws_base.cell(row=row_num, column=col_num).border = border

# Ajustar larguras das colunas
column_widths = [15, 20, 12, 12, 15, 18, 18, 20, 25, 15, 12]
for col_num, width in enumerate(column_widths, 1):
    ws_base.column_dimensions[ws_base.cell(row=1, column=col_num).column_letter].width = width

wb_base.save('data/base.xlsx')
print("✅ base.xlsx criado com sucesso!")

# ===== CRIAR ARQUIVO dashboard_xbox_finalizado.xlsx =====
print("📊 Criando dashboard_xbox_finalizado.xlsx...")

wb_dash = Workbook()

# Aba 1: Assets (Paleta de cores)
ws_assets = wb_dash.active
ws_assets.title = "Assets"
ws_assets['A1'] = "Paleta de Cores Xbox"
ws_assets['A1'].font = Font(bold=True, size=14)
ws_assets['A3'] = "Verde Principal"
ws_assets['B3'] = "#9BC848"
ws_assets['C3'].fill = PatternFill(start_color="9BC848", end_color="9BC848", fill_type="solid")
ws_assets['A4'] = "Verde Secundário"
ws_assets['B4'] = "#22C55E"
ws_assets['C4'].fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
ws_assets['A5'] = "Menu Principal"
ws_assets['B5'] = "#2AE6B1"
ws_assets['C5'].fill = PatternFill(start_color="2AE6B1", end_color="2AE6B1", fill_type="solid")
ws_assets['A6'] = "Menu Secundário"
ws_assets['B6'] = "#5BF6A8"
ws_assets['C6'].fill = PatternFill(start_color="5BF6A8", end_color="5BF6A8", fill_type="solid")

# Aba 2: Bases (mesma da base.xlsx)
ws_bases = wb_dash.create_sheet("Bases")
for col_num, header in enumerate(headers, 1):
    cell = ws_bases.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

for row_num, (idx, row) in enumerate(df.iterrows(), 2):
    ws_bases.cell(row=row_num, column=1, value=int(row['Subscriber ID']))
    ws_bases.cell(row=row_num, column=2, value=row['Name'])
    ws_bases.cell(row=row_num, column=3, value=row['Plan'])
    ws_bases.cell(row=row_num, column=4, value=row['Start Date'].strftime('%d/%m/%Y'))
    ws_bases.cell(row=row_num, column=5, value=row['Auto Renewal'])
    ws_bases.cell(row=row_num, column=6, value=row['Subscription Price'])
    ws_bases.cell(row=row_num, column=7, value=row['Subscription Type'])
    ws_bases.cell(row=row_num, column=8, value=row['EA Play Season Pass'])
    ws_bases.cell(row=row_num, column=9, value=row['Minecraft Season Pass'])
    ws_bases.cell(row=row_num, column=10, value=row['Coupon Value'])
    ws_bases.cell(row=row_num, column=11, value=row['Total Value'])

# Aba 3: Cálculos (análises)
ws_calc = wb_dash.create_sheet("Cálculos")
ws_calc['A1'] = "ANÁLISES E MÉTRICAS"
ws_calc['A1'].font = Font(bold=True, size=14)

# Análise 1: Faturamento por Tipo de Assinatura
ws_calc['A3'] = "1. Faturamento Total por Tipo de Assinatura"
ws_calc['A3'].font = Font(bold=True)
ws_calc['A4'] = "Tipo"
ws_calc['B4'] = "Faturamento Total"
ws_calc['C4'] = "Com Auto-Renovação"
ws_calc['D4'] = "Sem Auto-Renovação"

for col in ['A', 'B', 'C', 'D']:
    ws_calc[f'{col}4'].fill = header_fill
    ws_calc[f'{col}4'].font = header_font

row = 5
for sub_type in ['Monthly', 'Quarterly', 'Annual']:
    filtered = df[df['Subscription Type'] == sub_type]
    ws_calc[f'A{row}'] = sub_type
    ws_calc[f'B{row}'] = f"${filtered['Total Value'].sum():.2f}"
    ws_calc[f'C{row}'] = f"${filtered[filtered['Auto Renewal'] == 'Yes']['Total Value'].sum():.2f}"
    ws_calc[f'D{row}'] = f"${filtered[filtered['Auto Renewal'] == 'No']['Total Value'].sum():.2f}"
    row += 1

# Análise 2: Vendas de Add-ons
ws_calc['A8'] = "2. Vendas de Add-ons"
ws_calc['A8'].font = Font(bold=True)
ws_calc['A9'] = "Add-on"
ws_calc['B9'] = "Quantidade"
ws_calc['C9'] = "Faturamento Total"
ws_calc['A9'].fill = header_fill
ws_calc['B9'].fill = header_fill
ws_calc['C9'].fill = header_fill
ws_calc['A9'].font = header_font
ws_calc['B9'].font = header_font
ws_calc['C9'].font = header_font

ea_count = (df['EA Play Season Pass'] > 0).sum()
ea_total = df['EA Play Season Pass'].sum()
minecraft_count = (df['Minecraft Season Pass'] > 0).sum()
minecraft_total = df['Minecraft Season Pass'].sum()

ws_calc['A10'] = "EA Play Season Pass"
ws_calc['B10'] = ea_count
ws_calc['C10'] = f"${ea_total:.2f}"

ws_calc['A11'] = "Minecraft Season Pass"
ws_calc['B11'] = minecraft_count
ws_calc['C11'] = f"${minecraft_total:.2f}"

# Análise 3: Resumo Geral
ws_calc['A13'] = "3. Resumo Geral"
ws_calc['A13'].font = Font(bold=True)
ws_calc['A14'] = "Métrica"
ws_calc['B14'] = "Valor"
ws_calc['A14'].fill = header_fill
ws_calc['B14'].fill = header_fill
ws_calc['A14'].font = header_font
ws_calc['B14'].font = header_font

ws_calc['A15'] = "Total de Assinaturas"
ws_calc['B15'] = len(df)
ws_calc['A16'] = "Faturamento Total"
ws_calc['B16'] = f"${df['Total Value'].sum():.2f}"
ws_calc['A17'] = "Faturamento Médio"
ws_calc['B17'] = f"${df['Total Value'].mean():.2f}"
ws_calc['A18'] = "Total de Descontos Aplicados"
ws_calc['B18'] = f"${df['Coupon Value'].sum():.2f}"

# Aba 4: Dashboard (placeholder para visualizações)
ws_dashboard = wb_dash.create_sheet("Dashboard")
ws_dashboard['A1'] = "DASHBOARD XBOX GAME PASS"
ws_dashboard['A1'].font = Font(bold=True, size=16, color="9BC848")
ws_dashboard['A3'] = "Este é o espaço para criar suas visualizações e gráficos."
ws_dashboard['A4'] = "Use os dados da aba 'Cálculos' para criar tabelas dinâmicas e gráficos."

wb_dash.save('data/dashboard_xbox_finalizado.xlsx')
print("✅ dashboard_xbox_finalizado.xlsx criado com sucesso!")

# ===== ESTATÍSTICAS FINAIS =====
print("\n" + "="*60)
print("📊 ESTATÍSTICAS DOS DADOS GERADOS")
print("="*60)
print(f"Total de Registros: {len(df)}")
print(f"Faturamento Total: ${df['Total Value'].sum():.2f}")
print(f"Faturamento EA Play: ${df['EA Play Season Pass'].sum():.2f}")
print(f"Faturamento Minecraft: ${df['Minecraft Season Pass'].sum():.2f}")
print(f"Total de Descontos: ${df['Coupon Value'].sum():.2f}")
print("\nDistribuição por Plano:")
print(df['Plan'].value_counts())
print("\nDistribuição por Tipo de Assinatura:")
print(df['Subscription Type'].value_counts())
print("="*60)
print("✅ Arquivos criados com sucesso na pasta /data!")